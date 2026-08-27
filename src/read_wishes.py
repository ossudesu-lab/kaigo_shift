"""Excel帳票の色から希望を取り込む（設計メモ 3.3 / 6.4）

帳票では希望を**橙（薄茶）** で塗る運用になっている。その色のセルを拾って
`requests_YYYY-MM.csv` を書き出す。

使い方:
    python src/read_wishes.py 2026-10 --template "C:\\...\\シフト最新版3.xlsx" --sheet 10月

    --layout  行レイアウト名（既定は対象月の日数に合うもの。data/sheet_layout.yaml）
    --out     出力先（既定は requests_YYYY-MM.csv）
    --force   出力先が既にあっても上書きする

入→明→休 のセットが塗ってある場合は、**入の日だけ**を記録する。
明と休はハード制約C2/C3で自動的に続くため、希望として重ねて書く必要がない。

帳票の色の意味（2026年9月シートで確認）:
    橙（薄茶）  希望              ← これを拾う
    赤紫        当日リーダー
    淡桃        夜勤セット
    水色        検診・研修・有給   ← 件数だけ報告する（設計9章の保留事項）
"""

from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl

from excel_layout import (
    check_pool_rows, detect_day_cols, days_in, load_layout, pick_layout,
)

# 帳票記号 → 内部コード。write_excel.py の SYMBOL の逆引き。
CODE = {"Ａ": "DAY_L", "ＡC": "DAY_LATE", "Ｂ": "EARLY", "Ｃ": "LATE",
        "入": "NIGHT_IN", "明": "NIGHT_OUT", "休": "OFF", "有": "PAID"}

# 有給は「希望」ではなく申請済みの休み。色に関係なく拾い、fixed で記録する。
PAID_SYMBOL = "有"

# 橙（薄茶）＝希望。元の帳票はテーマ色9で塗られているが、write_excel.py が
# 書き戻すときは同じ色をRGBで指定する。見た目は同じでも別物として保存されるため、
# **両方を希望として認識する**。片方だけだと書き戻したシートを読み返せない。
THEME_ORANGE = 9
RGB_ORANGE = "FFF79646"     # テーマ色9（accent6）のRGB値
RGB_LIGHTBLUE = "FFCCECFF"  # 水色＝検診・研修


def fill_kind(cell) -> str | None:
    """セルの塗りを 'wish' / 'leave' / None で返す。"""
    f = cell.fill
    if f is None or f.patternType != "solid":
        return None
    fg = f.fgColor
    if fg.type == "theme" and fg.theme == THEME_ORANGE:
        return "wish"
    if fg.type == "rgb":
        if fg.rgb == RGB_ORANGE:
            return "wish"
        if fg.rgb == RGB_LIGHTBLUE:
            return "leave"
    return None


def extract(ws, layout: dict, year: int, month: int, ndays: int):
    """(希望リスト, 警告リスト, 有給らしきセル数) を返す。"""
    c0, _ = detect_day_cols(ws, ndays)

    # 行レイアウトの取り違えは黙って間違った希望を作るので、ここで止める
    # プール10名に加え、日勤のみ職員の行も読む対象なので、検証も両方にかける。
    all_rows = dict(layout["pool_rows"])
    all_rows.update(layout.get("daystaff_rows") or {})

    # 希望の記入用シートは空欄が正常。空欄そのものは問題にしない。
    problems = check_pool_rows(ws, layout, c0, ndays,
                               require_filled=False, rowmap=all_rows)
    if problems == ["__EMPTY_SHEET__"]:
        # 何も書かれていない白紙。エラーではなく「希望0件」として扱う。
        return [], ["シートに希望が1件も書かれていません（白紙のままのようです）"], 0
    if problems:
        raise SystemExit(
            "職員行として指定した行が、勤務表の行に見えません。\n"
            + "\n".join(f"  - {x}" for x in problems)
            + "\n  --layout が対象シートと合っているか確認してください"
              "（集計行を職員行と取り違えている可能性があります）。")

    rows, warns, leave = [], [], 0
    for sid, r in all_rows.items():
        cells = []
        for d in range(1, ndays + 1):
            c = ws.cell(row=r, column=c0 + d - 1)
            cells.append((d, str(c.value).strip() if c.value is not None else "",
                          fill_kind(c)))
        val = {d: v for d, v, _ in cells}
        kind = {d: k for d, _, k in cells}
        leave += sum(1 for _, _, k in cells if k == "leave")

        for d, v, k in cells:
            # 有給は「希望」ではなく申請済みの休み。色に関係なく拾い、
            # strength=fixed で記録する（公休とは別枠。設計3.5）。
            if v == PAID_SYMBOL:
                rows.append((sid, f"{year}-{month:02d}-{d:02d}", "PAID", v,
                             "fixed", "年次有給休暇（公休とは別枠）"))
                continue
            if k != "wish":
                continue
            if v == "":
                warns.append(f"{sid} {d}日: 橙で塗られていますが記号が空です。"
                             f"希望する勤務（休・入など）を書いてください")
                continue
            if v not in CODE:
                warns.append(f"{sid} {d}日: 未知の記号 {v!r} を飛ばしました")
                continue
            # 入→明→休 は入の日だけ記録する（明・休はハード制約で自動的に続く）
            if v == "明" and kind.get(d - 1) == "wish" and val.get(d - 1) == "入":
                continue
            if (v == "休" and kind.get(d - 1) == "wish" and val.get(d - 1) == "明"
                    and kind.get(d - 2) == "wish" and val.get(d - 2) == "入"):
                continue
            note = "入→明→休は入の日だけ記録（明・休はハード制約で自動）" if v == "入" else ""
            rows.append((sid, f"{year}-{month:02d}-{d:02d}", CODE[v], v, "wish", note))
    return sorted(rows, key=lambda x: (x[0], x[1])), warns, leave


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    opt = {a.split("=")[0]: (a.split("=", 1)[1] if "=" in a else True)
           for a in sys.argv[1:] if a.startswith("--")}
    if not args:
        print(__doc__)
        sys.exit(2)

    ym = args[0]
    if len(ym) != 7 or ym[4] != "-":
        print(f"対象月は YYYY-MM 形式で指定してください: {ym}")
        sys.exit(2)
    year, month = int(ym[:4]), int(ym[5:])

    root = Path(__file__).resolve().parent.parent
    data_dir = root / "data"
    layout_all = load_layout(data_dir)

    template = opt.get("--template")
    sheet = opt.get("--sheet")
    if not template or not sheet:
        print("--template（ブック）と --sheet（希望を塗ったシート名）を指定してください。")
        sys.exit(2)
    template = Path(template)
    if not template.exists():
        print(f"ブックが見つかりません: {template}")
        sys.exit(2)

    ndays = days_in(year, month)
    lay_name, layout = pick_layout(layout_all, opt.get("--layout"))

    wb = openpyxl.load_workbook(template)   # 色を見るので data_only にしない
    if sheet not in wb.sheetnames:
        print(f"シート {sheet!r} がありません。候補: {wb.sheetnames}")
        sys.exit(2)

    print("=" * 70)
    print(f"希望の取り込み: {year}年{month}月（{ndays}日）")
    print(f"  ブック: {template}  シート「{sheet}」  行レイアウト「{lay_name}」")
    print("=" * 70)

    rows, warns, leave = extract(wb[sheet], layout, year, month, ndays)

    if warns:
        print("\n■ 警告")
        for w in warns:
            print(f"  ⚠ {w}")

    from collections import Counter
    n_paid = sum(1 for r in rows if r[2] == "PAID")
    print(f"\n■ 希望 {len(rows) - n_paid}件"
          + (f"／ 有給 {n_paid}件（fixed）" if n_paid else ""))
    print("  職員別:", dict(sorted(Counter(s for s, *_ in rows).items())) or "なし")
    print("  種類別:", dict(Counter(c for _, _, c, *_ in rows)) or "なし")
    if leave:
        print(f"\n※ 水色（検診・研修）のセルが {leave}件 あります。"
              "\n  有給は記号「有」で判定するので、色は見ていません。")

    out = Path(opt.get("--out") or (root / f"requests_{year}-{month:02d}.csv"))
    if out.exists() and "--force" not in opt:
        print(f"\n既にあります: {out}\n上書きするなら --force を付けてください。")
        sys.exit(1)
    with out.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        w.writerow(["staff_id", "date", "allowed", "symbol", "strength", "note"])
        w.writerows(rows)
    print(f"\n出力: {out}")


if __name__ == "__main__":
    main()
