"""求解結果を Excel の帳票に書き戻す（設計メモ 6.4 / 実装順序6）

**元のブックには絶対に書き込まない。** ファイルをコピーして別ブックを作り、
そこにだけ書く。元ファイルは図形・印刷設定・結合セルを含む手作りの帳票で、
Python から開いて保存し直すと崩れる恐れがあるため。

使い方:
    python src/write_excel.py 2026-10 --template "C:\\...\\シフト最新版3.xlsx"

    --sheet     雛形にするシート名（対象月と同じ日数のシートを指定する）
    --layout    行レイアウト名（既定は data/sheet_layout.yaml の default_layout）
    --out       出力先（既定は out/シフト_YYYY-MM.xlsx）
    --from-csv  求解し直さず、既存のCSVを読んで書き戻す
    --blank     求解せず、日付と曜日だけ入れた白紙シートを出す（希望の記入用）
    --force     出力先が既にあっても上書きする

書き込むのは変則プール10名の行だけ。プール外（管理者・相談員・看護・
非常勤・指導員・栄養士・送迎）はソルバーが決めないので、雛形の値を消して
空欄にする。そこは人が埋める。

色は帳票の慣習に合わせる（設計6.4）:
    赤紫  当日リーダー（Ａ／ＡC）
    淡桃  夜勤セット（入→明→休）
    橙    通った希望
"""

from __future__ import annotations

import shutil
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from excel_layout import check_pool_rows, detect_day_cols, load_layout, pick_layout
from fill_daystaff import solve_daystaff
from shift_solver import (
    DAY_L, DAY_LATE, NIGHT_IN, NIGHT_OUT, NONWORK, OFF, SYMBOL, WEEKDAY_JP,
    load_inputs, solve,
)

DAY_COL0 = 7  # G列 = 1日

FILL_LEADER = PatternFill("solid", fgColor="FFFD11CA")   # 赤紫: 当日リーダー
FILL_NIGHT = PatternFill("solid", fgColor="FFFFCCFF")    # 淡桃: 夜勤セット
FILL_WISH = PatternFill("solid", fgColor="FFF79646")     # 橙:   通った希望
FILL_NONE = PatternFill(fill_type=None)

# 集計列（日付列の右に7列）。原本の数式は参照範囲が月によってずれていたので、
# 対象月の範囲で書き直す。
TALLY = [("日勤", "Ａ"), ("早出", "Ｂ"), ("遅出", "Ｃ"), ("遅２", "D"),
         ("夜勤", "入"), ("夜勤", "明"), ("休", "休")]


def night_set_days(assign: dict, sid: str, ND: int) -> set[int]:
    """夜勤セット（入・明・明の翌日の休）に当たる日を返す。"""
    out = set()
    for d in range(ND):
        code = assign.get((sid, d))
        if code in (NIGHT_IN, NIGHT_OUT):
            out.add(d)
        elif code in NONWORK and d >= 1 and assign.get((sid, d - 1)) == NIGHT_OUT:
            out.add(d)
    return out


def granted_wish_days(inp, assign: dict) -> set[tuple[str, int]]:
    """通った希望のセル。希望どおりの勤務が入った日だけを対象にする。"""
    days = {dt: i for i, dt in enumerate(inp.days)}
    out = set()
    for r in inp.requests:
        d = days.get(r.date)
        if d is not None and assign.get((r.staff_id, d)) in r.allowed:
            out.add((r.staff_id, d))
    return out


def write_sheet(ws, inp, assign: dict, layout: dict, blank: bool = False) -> list[str]:
    """雛形シートを対象月の内容で塗り替える。注意点のリストを返す。
    blank=True なら勤務欄を空にして、日付と曜日だけを入れる（希望の記入用）。"""
    notes = []
    ND = len(inp.days)
    c0, c1 = detect_day_cols(ws, ND)

    # 書き込む前に、指定した行が本当に職員行かを確かめる。
    # 取り違えたまま書くと集計行を潰してしまう。
    problems = check_pool_rows(ws, layout, c0, ND)
    if problems:
        raise SystemExit(
            "職員行として指定した行が、勤務表の行に見えません。書き込みを中止します。\n"
            + "\n".join(f"  - {x}" for x in problems)
            + "\n  --sheet が対象シートと合っているか確認してください。")

    ws["A2"] = f"{inp.month}月勤務表"
    for d in range(ND):
        ws.cell(row=2, column=c0 + d).value = d + 1
        ws.cell(row=4, column=c0 + d).value = WEEKDAY_JP[inp.days[d].weekday()]

    if blank:
        # 希望の記入用。全職員の勤務欄を空にして、日付と曜日だけを残す。
        for r in list(layout["pool_rows"].values()) + layout["other_rows"]:
            for d in range(ND):
                c = ws.cell(row=r, column=c0 + d)
                c.value = None
                c.fill = FILL_NONE
        notes.append("全職員の勤務欄を空にしました。希望を書いて薄茶色に塗ってください")
        notes.append("有給は記号「有」で書いてください（色は不要・公休とは別枠）")
        return notes

    # プール外の行は雛形の値が残るので消す（ソルバーは決めない）
    for r in layout["other_rows"]:
        for d in range(ND):
            ws.cell(row=r, column=c0 + d).value = None
            ws.cell(row=r, column=c0 + d).fill = FILL_NONE
    if layout["other_rows"]:
        notes.append(f"指導員・栄養士・送迎 {len(layout['other_rows'])}行を空欄にしました。"
                     "人が埋めてください")

    wishes = granted_wish_days(inp, assign)
    rowmap = dict(layout["pool_rows"])
    rowmap.update(layout.get("daystaff_rows") or {})
    for sid, row in rowmap.items():
        if not any((sid, d) in assign for d in range(ND)):
            continue   # 割当が無い職員（雛形に行はあるが対象外）は触らない
        nights = night_set_days(assign, sid, ND)
        for d in range(ND):
            code = assign.get((sid, d))
            cell = ws.cell(row=row, column=c0 + d)
            cell.value = SYMBOL.get(code) if code else None
            if (sid, d) in wishes:
                cell.fill = FILL_WISH
            elif code in (DAY_L, DAY_LATE):
                cell.fill = FILL_LEADER
            elif d in nights:
                cell.fill = FILL_NIGHT
            else:
                cell.fill = FILL_NONE
        # 集計列（原本の数式は参照範囲がずれていたので対象月の範囲で書き直す）
        a, b = get_column_letter(c0), get_column_letter(c1)
        for i, (_, sym) in enumerate(TALLY):
            ws.cell(row=row, column=c1 + 1 + i).value = \
                f'=COUNTIF({a}{row}:{b}{row},"{sym}")'

    return notes


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("--")]
    opt = {a.split("=")[0]: (a.split("=", 1)[1] if "=" in a else True)
           for a in sys.argv[1:] if a.startswith("--")}
    if not args:
        print(__doc__)
        sys.exit(2)

    ym = args[0]
    if len(ym) != 7 or ym[4] != "-":
        print(f"対象月は YYYY-MM 形式で指定してください: {ym}")
        sys.exit(2)
    year, month = int(ym[:4]), int(ym[5:])

    root = Path(__file__).resolve().parent.parent
    data_dir = root / "data"
    layout_all = load_layout(data_dir)

    template = opt.get("--template")
    if not template:
        print("--template で雛形にするブックを指定してください（元ファイルは変更しません）。")
        sys.exit(2)
    template = Path(template)
    if not template.exists():
        print(f"雛形が見つかりません: {template}")
        sys.exit(2)

    inp = load_inputs(data_dir, (year, month))
    ND = len(inp.days)
    # 雛形シートと行レイアウトは別物。シートは日数が合うものを、
    # レイアウトは職員の並びが合うものを選ぶ（設計6.4）。
    lay_name, layout = pick_layout(layout_all, opt.get("--layout"))
    sheet = opt.get("--sheet")
    if not sheet:
        print("--sheet で雛形にするシート名を指定してください"
              f"（対象月と同じ{ND}日ぶんの列があるもの）。")
        sys.exit(2)

    out = Path(opt.get("--out") or (root / "out" / f"シフト_{year}-{month:02d}.xlsx"))
    if out.resolve() == template.resolve():
        print("出力先が雛形と同じです。元ファイルには書き込みません。")
        sys.exit(2)
    if out.exists() and "--force" not in opt:
        print(f"既にあります: {out}\n上書きするなら --force を付けてください。")
        sys.exit(1)

    print("=" * 70)
    print(f"Excelへの書き戻し: {year}年{month}月（{ND}日）")
    print(f"  雛形: {template}  シート「{sheet}」  行レイアウト「{lay_name}」")
    print(f"  出力: {out}")
    print("=" * 70)

    from_csv = opt.get("--from-csv")
    blank = "--blank" in opt
    mgr_days: set[int] = set()
    if blank:
        assign = {}
        print()
        print("白紙シートを作ります（求解しません）。")
    elif from_csv:
        from validate_shift import read_edited_csv
        assign, _, _, blanks, errors, _ = read_edited_csv(Path(from_csv), inp)
        if errors:
            print("\n■ 読み込みエラー")
            for e in errors:
                print(f"  - {e}")
            sys.exit(1)
        print()
        print(f"CSVから読み込みました: {from_csv}"
              + (f"（未割当 {len(blanks)}件）" if blanks else ""))
        # CSVには管理者のリーダー日が入っていないので、プール側から逆算する。
        # プールの誰もリーダーを担っていない日 = 管理者が入った日。
        pool_ids = [s.staff_id for s in inp.staff.values() if s.in_pool]
        mgr_days = {d for d in range(len(inp.days))
                    if not any(assign.get((s, d)) in (DAY_L, DAY_LATE) for s in pool_ids)}
    else:
        print("\n求解中…（月によっては数分かかります）")
        res = solve(inp)
        if not res["feasible"]:
            print(f"解が見つかりませんでした（{res['status']}）。")
            sys.exit(1)
        assign = res["assign"]
        print(f"  {res['status']}  目的値={int(res['objective'])}  "
              f"再現性={'あり' if res.get('canonical') else 'なし'}")
        mgr_days = {(d - inp.days[0]).days for d in res["mgr_days"]}
        if res["mgr_days"]:
            print(f"  管理者がリーダーを担う日: "
                  + "、".join(f"{d.day}日" for d in res["mgr_days"]))

    # 日勤のみ職員（管理者・相談員・看護・パート）を組む。プール側とは
    # 枠を奪い合わないので独立して解ける（設計5.6）。
    if not blank:
        mgr_id = inp.manager_id
        day_assign = solve_daystaff(
            inp, must_work={mgr_id: mgr_days} if mgr_id else None)
        if day_assign:
            assign = {**assign, **day_assign}
        else:
            print("  ※ 日勤のみ職員は組めませんでした。その行は空欄のままにします。")

    # 元ファイルはコピーしてから触る。ここから先は out だけを操作する。
    out.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(template, out)
    wb = openpyxl.load_workbook(out)
    ws = wb.copy_worksheet(wb[sheet])
    # 先に他シートを消してからリネームする。雛形が同名だと openpyxl が
    # 重複を避けて「9月1」のような名前に変えてしまうため。
    for name in list(wb.sheetnames):
        if name != ws.title:
            del wb[name]
    ws.title = f"{month}月"

    notes = write_sheet(ws, inp, assign, layout, blank=blank)
    wb.save(out)

    print(f"\n■ 書き込みました: シート「{ws.title}」")
    for n in notes:
        print(f"  ・{n}")
    print("  ・色: 赤紫=当日リーダー / 淡桃=夜勤セット / 橙=通った希望")
    print(f"\n出力: {out}")


if __name__ == "__main__":
    main()
